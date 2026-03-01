import sys
import asyncio

if sys.platform.startswith("win"):
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

import os
import io
import csv
import json
import re
import secrets
import datetime
from pathlib import Path
from typing import Dict, Any, Optional, Tuple, List
import anyio

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.requests import Request
from starlette.responses import Response

import fitz  # PyMuPDF
import pdfplumber

# Gemini (google-genai) – same import style as your web_app.py
try:
    from google import genai
    from google.genai import types
except Exception:
    genai = None
    types = None

APP_DIR = os.path.dirname(os.path.abspath(__file__))

app = FastAPI(title="DBS Check")
app.mount("/static", StaticFiles(directory=os.path.join(APP_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(APP_DIR, "templates"))

# -------------------------
# Lightweight session (cookie) for per-session run lock
# -------------------------
SID_COOKIE = "sid"


def _new_sid() -> str:
    # URL-safe random id
    return secrets.token_urlsafe(18)


@app.middleware("http")
async def _session_middleware(request: Request, call_next):
    _ensure_cleanup_task()

    sid = request.cookies.get(SID_COOKIE)
    if not sid:
        sid = _new_sid()
        request.state.sid = sid
        response: Response = await call_next(request)
        response.set_cookie(
            SID_COOKIE,
            sid,
            httponly=True,
            samesite="lax",
            secure=False,
            max_age=60 * 60 * 24 * 7,  # 7 days
        )
        return response

    request.state.sid = sid
    response: Response = await call_next(request)
    return response


def _sid_from(request: Request) -> str:
    sid = getattr(request.state, "sid", None)
    if sid:
        return sid
    return request.cookies.get(SID_COOKIE) or _new_sid()



def _env(name: str, default: Optional[str] = None) -> Optional[str]:
    v = os.getenv(name)
    if v is None or str(v).strip() == "":
        return default
    return v


# --- Env (FAST -> STRONG fallback) ---
GEMINI_API_KEY = _env("GEMINI_API_KEY")
GEMINI_MODEL_FAST = _env("GEMINI_MODEL_FAST", "gemini-2.0-flash-001")
GEMINI_MODEL_STRONG = _env("GEMINI_MODEL_STRONG", "gemini-2.5-pro")


def get_gemini_client():
    if not GEMINI_API_KEY or genai is None or types is None:
        return None
    try:
        return genai.Client(api_key=GEMINI_API_KEY)
    except Exception:
        return None


GEMINI_CLIENT = get_gemini_client()


def normalize_ws(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip())


MONTHS = {
    "JANUARY": "01", "FEBRUARY": "02", "MARCH": "03", "APRIL": "04", "MAY": "05", "JUNE": "06",
    "JULY": "07", "AUGUST": "08", "SEPTEMBER": "09", "OCTOBER": "10", "NOVEMBER": "11", "DECEMBER": "12",
    "JAN": "01", "FEB": "02", "MAR": "03", "APR": "04", "JUN": "06", "JUL": "07", "AUG": "08",
    "SEP": "09", "SEPT": "09", "OCT": "10", "NOV": "11", "DEC": "12",
}


def parse_uk_date_words(date_str: str) -> Optional[Tuple[str, str, str]]:
    """Parse '11 SEPTEMBER 2023' -> ('11','09','2023')"""
    if not date_str:
        return None
    s = normalize_ws(date_str).upper()
    m = re.search(r"\b(\d{1,2})\s+([A-Z]{3,9})\s+(\d{4})\b", s)
    if not m:
        return None
    dd = m.group(1).zfill(2)
    mon = MONTHS.get(m.group(2))
    yyyy = m.group(3)
    if not mon:
        return None
    return dd, mon, yyyy


def parse_ddmmyyyy(date_str: str) -> Optional[Tuple[str, str, str]]:
    if not date_str:
        return None
    s = normalize_ws(date_str)
    m = re.search(r"\b(\d{1,2})[\/\.-](\d{1,2})[\/\.-](\d{2,4})\b", s)
    if not m:
        return None
    dd = m.group(1).zfill(2)
    mm = m.group(2).zfill(2)
    yy = m.group(3)
    if len(yy) == 2:
        yy = ("19" if int(yy) > 30 else "20") + yy
    return dd, mm, yy


def validate_cert_number(s: str) -> Optional[str]:
    if not s:
        return None
    digits = re.sub(r"\D", "", s)
    if 10 <= len(digits) <= 14:
        return digits
    return None


def score_cert_number(val: str) -> int:
    digits = re.sub(r"\D", "", val or "")
    if not digits:
        return 0
    # Base score for plausible length
    if 10 <= len(digits) <= 14:
        base = 80
    else:
        return 30
    # Extra points if it's already validated by our validator (currently length-based, but future-proof)
    return min(100, base + (20 if validate_cert_number(digits) else 0))


def score_surname(val: str, *, source: str = "") -> int:
    s = (val or "").strip()
    if not s:
        return 0
    # Allow letters, spaces and hyphens only
    cleaned = re.sub(r"[^A-Za-z\-\s]", "", s).strip()
    if not cleaned:
        return 10
    length = len(cleaned.replace(" ", ""))
    if length < 2:
        score = 35
    elif length < 4:
        score = 70
    else:
        score = 85
    if cleaned != s:
        score -= 10
    # Minor adjustment based on extraction source (PDF text tends to be cleaner)
    if (source or "").lower().startswith("pdf"):
        score = min(100, score + 5)
    return max(0, min(100, score))


def score_dob(dd: str, mm: str, yyyy: str, *, source: str = "") -> int:
    dd = (dd or "").strip()
    mm = (mm or "").strip()
    yyyy = (yyyy or "").strip()
    if not (dd or mm or yyyy):
        return 0
    # Partial date
    if not (dd and mm and yyyy):
        return 45
    try:
        d = int(dd); m = int(mm); y = int(yyyy)
        datetime.date(y, m, d)
        score = 95
        if (source or "").lower().startswith("pdf"):
            score = min(100, score + 5)
        return score
    except Exception:
        return 20


def overall_confidence(cert_score: int, surname_score: int, dob_score: int) -> int:
    scores = [s for s in [cert_score, surname_score, dob_score] if s > 0]
    if not scores:
        return 0
    avg = sum(scores) / len(scores)
    # Penalty if something is missing
    missing = 3 - len(scores)
    avg = avg - (missing * 8)
    return max(0, min(100, int(round(avg))))


# NOTE: The product no longer surfaces an "overall" confidence score in the UI.
# We keep this helper only for backward compatibility, but responses should
# include per-field confidence only.


# -------------------------
# PDF/Text extraction

# -------------------------
def extract_text_from_pdf(pdf_bytes: bytes, max_pages: int = 2) -> str:
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            texts = []
            for page in pdf.pages[:max_pages]:
                t = page.extract_text() or ""
                if t.strip():
                    texts.append(t)
            return "\n".join(texts).strip()
    except Exception:
        return ""


def pdf_to_images_bytes(pdf_bytes: bytes, max_pages: int = 1, dpi: int = 240) -> List[bytes]:
    """
    Render the first page (and a few useful crops) to PNG bytes.

    This is used for scanned PDFs before sending to Gemini Vision.
    Hardened for production (bad uploads, encrypted PDFs) to avoid 500s on Render.
    """
    images: List[bytes] = []

    # Basic sanity checks
    if not pdf_bytes or len(pdf_bytes) < 100:
        raise ValueError("Uploaded file is empty or too small to be a valid PDF.")
    if not pdf_bytes.lstrip().startswith(b"%PDF"):
        raise ValueError("Uploaded file does not look like a PDF (missing %PDF header).")

    # Open PDF
    try:
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as e:
        raise ValueError(f"PyMuPDF failed to open PDF (corrupt/unsupported/encrypted). Details: {e}")

    try:
        # Encrypted PDFs
        if getattr(doc, "is_encrypted", False):
            try:
                doc.authenticate("")  # try empty password
            except Exception:
                pass
            if getattr(doc, "is_encrypted", False):
                raise ValueError("PDF is password-protected. Please upload an unlocked PDF.")

        if getattr(doc, "page_count", 0) <= 0:
            return images

        page = doc.load_page(0)
        zoom = dpi / 72.0
        mat = fitz.Matrix(zoom, zoom)

        rect = page.rect
        # Broad bands (work well across most DBS layouts)
        bands = [
            (0.0, 0.0, 1.0, 0.32),   # header/top
            (0.0, 0.28, 1.0, 0.72),  # middle (name + DOB often here)
            (0.0, 0.68, 1.0, 1.0),   # bottom (dates/notes)
        ]

        # Crops first
        for x0, y0, x1, y1 in bands:
            clip = fitz.Rect(
                rect.x0 + rect.width * x0,
                rect.y0 + rect.height * y0,
                rect.x0 + rect.width * x1,
                rect.y0 + rect.height * y1,
            )
            pix = page.get_pixmap(matrix=mat, clip=clip, alpha=False)
            images.append(pix.tobytes("png"))

        # Full page last
        pix_full = page.get_pixmap(matrix=mat, alpha=False)
        images.append(pix_full.tobytes("png"))

        return images
    finally:
        try:
            doc.close()
        except Exception:
            pass

# -------------------------
# Regex extraction (only works if PDF has a text layer)
# -------------------------
def extract_fields_from_text(text: str) -> Dict[str, Any]:
    t = normalize_ws(text)
    out: Dict[str, Any] = {"certificate_number": None, "surname": None, "dob": None}

    m = re.search(r"Certificate\s*Number[:\s]*([0-9\s]{8,20})", t, flags=re.IGNORECASE)
    if m:
        out["certificate_number"] = validate_cert_number(m.group(1))

    m = re.search(r"Surname[:\s]*([A-Z'\-\s]{2,40})", t, flags=re.IGNORECASE)
    if m:
        out["surname"] = normalize_ws(m.group(1)).upper()

    m = re.search(
        r"Date\s*of\s*Birth[:\s]*([0-9]{1,2}\s+[A-Za-z]{3,9}\s+[0-9]{4}|\d{1,2}[\/\.-]\d{1,2}[\/\.-]\d{2,4})",
        t,
        flags=re.IGNORECASE,
    )
    if m:
        parts = parse_uk_date_words(m.group(1)) or parse_ddmmyyyy(m.group(1))
        if parts:
            out["dob"] = {"dd": parts[0], "mm": parts[1], "yyyy": parts[2]}

    return out


# -------------------------
# Gemini Vision extraction (FAST -> STRONG fallback)
# -------------------------
VISION_PROMPT = """You are extracting ONLY these fields from a UK DBS Enhanced Certificate (page 1):
1) Certificate Number (digits only)
2) Applicant Surname (value next to label 'Surname')
3) Applicant Forename(s) (value next to label 'Forename(s)' or 'Forenames')
4) Date of Birth (label 'Date of Birth')
5) Date of Issue (label 'Date of Issue' or 'Issue Date') — DO NOT use Print Date / Date Printed.

Return STRICT JSON with this schema (confidence must be 0.0 to 1.0):
{
  "certificate_number": {"value": "string digits or empty", "confidence": 0.0},
  "surname": {"value": "string or empty", "confidence": 0.0},
  "forename": {"value": "string or empty", "confidence": 0.0},
  "dob": {"day": "DD or empty", "month": "MM or empty", "year": "YYYY or empty", "confidence": 0.0},
  "issue_date": {"day": "DD or empty", "month": "MM or empty", "year": "YYYY or empty", "confidence": 0.0}
}

Rules:
- If a field is not visible, return empty value(s) and confidence 0.0.
- Certificate number MUST be digits only (no spaces).
- Do not guess. If uncertain, lower confidence.
- Never return 1.0 unless the field is perfectly clear in the document.
"""


def _parse_json_response(text: str) -> Dict[str, Any]:
    if not text:
        return {}
    text = text.strip()
    text = re.sub(r"^```(json)?", "", text, flags=re.IGNORECASE).strip()
    text = re.sub(r"```$", "", text).strip()
    m = re.search(r"\{.*\}", text, flags=re.DOTALL)
    if not m:
        return {}
    try:
        return json.loads(m.group(0))
    except Exception:
        return {}


def gemini_vision_extract_images(images: List[Tuple[bytes, str]]) -> Dict[str, Any]:
    """Vision extraction with FAST -> STRONG fallback.

    Returns:
      {
        certificate_number: str|None,
        surname: str|None,
        dob: {dd,mm,yyyy}|None,
        _model: str,
      }
    """
    if GEMINI_CLIENT is None or not images:
        return {}

    def _call(model_name: str) -> Dict[str, Any]:
        parts = [types.Part.from_text(text=VISION_PROMPT)]
        for b, mime in images:
            parts.append(types.Part.from_bytes(data=b, mime_type=mime))
        resp = GEMINI_CLIENT.models.generate_content(
            model=model_name,
            contents=[types.Content(role="user", parts=parts)],
        )
        txt = getattr(resp, "text", None) or ""
        data = _parse_json_response(txt)
        if isinstance(data, dict):
            data["_model"] = model_name
        return data if isinstance(data, dict) else {}

    # 1) FAST first
    data: Dict[str, Any] = {}
    try:
        data = _call(GEMINI_MODEL_FAST)
    except Exception:
        data = {}

    # Decide if we need STRONG (any key missing)
    def _is_missing(d: Dict[str, Any]) -> bool:
        if not d:
            return True
        def _val(x):
            if isinstance(x, dict):
                return str(x.get("value") or "")
            return str(x or "")
        cert_ok = bool(validate_cert_number(_val(d.get("certificate_number", ""))))
        surname_ok = bool(normalize_ws(_val(d.get("surname", ""))).strip())
        dob_v = d.get("dob")
        if isinstance(dob_v, dict):
            dob_ok = bool(str(dob_v.get("day") or "").strip() and str(dob_v.get("month") or "").strip() and str(dob_v.get("year") or "").strip())
        else:
            dob_ok = bool(normalize_ws(str(dob_v or "")).strip())
        return not (cert_ok and surname_ok and dob_ok)

    if _is_missing(data):
        try:
            data = _call(GEMINI_MODEL_STRONG)
        except Exception:
            pass

    model_used = str(data.get("_model") or (GEMINI_MODEL_STRONG if _is_missing(data) else GEMINI_MODEL_FAST))

    def _get_val(name: str) -> str:
        v = data.get(name)
        if isinstance(v, dict):
            return str(v.get("value") or "")
        return str(v or "")

    def _get_conf(name: str) -> float:
        v = data.get(name)
        if isinstance(v, dict):
            try:
                return float(v.get("confidence") or 0.0)
            except Exception:
                return 0.0
        return 0.0

    def _pct(f: float) -> int:
        try:
            f = float(f or 0.0)
        except Exception:
            f = 0.0
        f = max(0.0, min(1.0, f))
        return int(round(f * 100))

    out: Dict[str, Any] = {
        "certificate_number": validate_cert_number(_get_val("certificate_number")),
        "surname": normalize_ws(_get_val("surname")).upper() or None,
        "forename": normalize_ws(_get_val("forename")).upper() or None,
        "dob": None,
        "issue_date": None,
        "confidence": {
            "certificate_number": _pct(_get_conf("certificate_number")),
            "surname": _pct(_get_conf("surname")),
            "dob": _pct(_get_conf("dob")),
            "issue_date": _pct(_get_conf("issue_date")),
        },
        "_model": model_used,
    }

    # DOB may arrive as structured dict
    dob_obj = data.get("dob") if isinstance(data.get("dob"), dict) else None
    if dob_obj:
        dd = str(dob_obj.get("day") or "").zfill(2) if str(dob_obj.get("day") or "").strip() else ""
        mm = str(dob_obj.get("month") or "").zfill(2) if str(dob_obj.get("month") or "").strip() else ""
        yy = str(dob_obj.get("year") or "").strip()
        if dd and mm and yy:
            out["dob"] = {"dd": dd, "mm": mm, "yyyy": yy}
    else:
        dob_parts = (
            parse_uk_date_words(_get_val("dob"))
            or parse_ddmmyyyy(_get_val("dob"))
            or parse_uk_date_words(str(data.get("date_of_birth", "") or ""))
            or parse_ddmmyyyy(str(data.get("date_of_birth", "") or ""))
        )
        if dob_parts:
            out["dob"] = {"dd": dob_parts[0], "mm": dob_parts[1], "yyyy": dob_parts[2]}

    issue_obj = data.get("issue_date") if isinstance(data.get("issue_date"), dict) else None
    if issue_obj:
        dd = str(issue_obj.get("day") or "").zfill(2) if str(issue_obj.get("day") or "").strip() else ""
        mm = str(issue_obj.get("month") or "").zfill(2) if str(issue_obj.get("month") or "").strip() else ""
        yy = str(issue_obj.get("year") or "").strip()
        if dd and mm and yy:
            out["issue_date"] = {"dd": dd, "mm": mm, "yyyy": yy}
    else:
        issue_parts = (
            parse_uk_date_words(_get_val("issue_date"))
            or parse_ddmmyyyy(_get_val("issue_date"))
            or parse_uk_date_words(str(data.get("date_of_issue", "") or ""))
            or parse_ddmmyyyy(str(data.get("date_of_issue", "") or ""))
            or parse_uk_date_words(str(data.get("issued_on", "") or ""))
            or parse_ddmmyyyy(str(data.get("issued_on", "") or ""))
        )
        if issue_parts:
            out["issue_date"] = {"dd": issue_parts[0], "mm": issue_parts[1], "yyyy": issue_parts[2]}

    return out


# -------------------------
# Routes
# -------------------------
@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/health")
def health():
    return {
        "gemini_key_present": bool(GEMINI_API_KEY),
        "gemini_import_ok": bool(genai is not None and types is not None),
        "gemini_client_ok": bool(GEMINI_CLIENT is not None),
        "model_fast": GEMINI_MODEL_FAST,
        "model_strong": GEMINI_MODEL_STRONG,
    }



# -------------------------
# Job storage (ephemeral, Free-plan safe)
# -------------------------
import uuid
import zipfile
import time
from zoneinfo import ZoneInfo
from starlette.background import BackgroundTask

JOBS_ROOT = Path("/tmp") / "dbs_jobs"
JOBS_ROOT.mkdir(parents=True, exist_ok=True)

# job_id -> metadata/state
JOBS: Dict[str, Dict[str, Any]] = {}

# session_id -> active job_id (run lock: only 1 active job per session)
ACTIVE_BY_SID: Dict[str, str] = {}

# v2 spec: delete after 15 minutes inactivity (bulk and single)
JOB_INACTIVITY_TTL_SECONDS = 15 * 60

_CLEANUP_TASK_STARTED = False


def _now() -> float:
    return time.time()


def _touch_job(job_id: str) -> None:
    meta = JOBS.get(job_id)
    if meta:
        meta["last_access"] = _now()


def _delete_job(job_id: str) -> None:
    meta = JOBS.pop(job_id, None)
    if not meta:
        return
    # release session lock if held
    sid = meta.get("sid")
    if sid and ACTIVE_BY_SID.get(sid) == job_id:
        ACTIVE_BY_SID.pop(sid, None)

    try:
        job_path = Path(meta.get("path") or "")
        if job_path.exists():
            import shutil
            shutil.rmtree(job_path, ignore_errors=True)
    except Exception:
        pass




def _release_lock(job_id: str) -> None:
    meta = JOBS.get(job_id)
    if not meta:
        return
    sid = meta.get("sid")
    if sid and ACTIVE_BY_SID.get(sid) == job_id:
        ACTIVE_BY_SID.pop(sid, None)
def _cleanup_jobs_once() -> None:
    now = _now()
    for jid, meta in list(JOBS.items()):
        last_access = float(meta.get("last_access") or meta.get("created") or 0)
        mode = meta.get("mode") or ""
        zip_downloaded = bool(meta.get("zip_downloaded"))
        # Bulk: if zip downloaded -> delete immediately
        if mode == "bulk" and zip_downloaded:
            _delete_job(jid)
            continue
        # Otherwise: inactivity TTL
        if now - last_access > JOB_INACTIVITY_TTL_SECONDS:
            _delete_job(jid)


async def _cleanup_loop() -> None:
    while True:
        try:
            _cleanup_jobs_once()
        except Exception:
            pass
        await anyio.sleep(60)


def _ensure_cleanup_task() -> None:
    global _CLEANUP_TASK_STARTED
    if _CLEANUP_TASK_STARTED:
        return
    _CLEANUP_TASK_STARTED = True
    try:
        import asyncio
        asyncio.get_event_loop().create_task(_cleanup_loop())
    except Exception:
        # If event loop isn't ready, FastAPI startup will call again.
        _CLEANUP_TASK_STARTED = False


def _new_job_dir(*, prefix: str, sid: str, mode: str) -> Tuple[str, Path]:
    _cleanup_jobs_once()
    job_id = str(uuid.uuid4())
    job_dir = JOBS_ROOT / f"{prefix}-{job_id}"
    job_dir.mkdir(parents=True, exist_ok=True)
    JOBS[job_id] = {
        "path": str(job_dir),
        "created": _now(),
        "last_access": _now(),
        "sid": sid,
        "mode": mode,  # "single" or "bulk"
        "state": "running",
        "checked_date": "",
        "rows": [],          # bulk rows
        "zip_name": "",
        "zip_ready": False,
        "zip_downloaded": False,
        "message": "",
    }
    # Run lock
    ACTIVE_BY_SID[sid] = job_id
    return job_id, job_dir


def _uk_checked_date() -> str:
    # "Checked date" based on UK time.
    dt = datetime.datetime.now(tz=ZoneInfo("Europe/London"))
    return dt.strftime("%d.%m.%Y")


def _safe_filename(name: str, default: str) -> str:
    name = normalize_ws(name).strip()
    if not name:
        name = default
    # Replace illegal filename chars (Windows + URL safety)
    name = re.sub(r'[\\/:*?"<>|]+', "-", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name



def _uk_checked_date() -> str:
    # "Checked date" based on UK time.
    dt = datetime.datetime.now(tz=ZoneInfo("Europe/London"))
    return dt.strftime("%d.%m.%Y")


def _safe_filename(name: str, default: str) -> str:
    name = normalize_ws(name).strip()
    if not name:
        name = default
    # Replace illegal filename chars (Windows + URL safety)
    name = re.sub(r'[\\/:*?"<>|]+', "-", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name



# -------------------------
# Extract (single or bulk)
# ---------------
# -------------------------
# Spreadsheet parsing (CSV/XLSX) for bulk mode
# -------------------------
def _norm_col(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (name or "").strip().lower())

def _parse_dob_value(v: Any) -> Tuple[str, str, str]:
    """Return (dd, mm, yyyy) as strings from a variety of common DOB formats."""
    if v is None:
        return ("", "", "")
    # openpyxl may return datetime/date
    if isinstance(v, (datetime.date, datetime.datetime)):
        d = v.date() if isinstance(v, datetime.datetime) else v
        return (str(d.day).zfill(2), str(d.month).zfill(2), str(d.year))
    s = str(v).strip()
    if not s:
        return ("", "", "")
    # Accept YYYY-MM-DD
    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", s)
    if m:
        yy, mm, dd = m.group(1), m.group(2), m.group(3)
        return (dd.zfill(2), mm.zfill(2), yy)
    # Accept DD/MM/YYYY or DD-MM-YYYY
    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", s)
    if m:
        dd, mm, yy = m.group(1), m.group(2), m.group(3)
        return (dd.zfill(2), mm.zfill(2), yy)
    return ("", "", "")

def parse_csv_rows(content: bytes) -> List[Dict[str, Any]]:
    # Try utf-8-sig first, then latin-1
    text = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(enc)
            break
        except Exception:
            continue
    if text is None:
        raise ValueError("Unable to read CSV encoding.")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("CSV has no header row.")
    cols = { _norm_col(c): c for c in reader.fieldnames }
    return _rows_from_dict_iter(reader, cols)

def parse_xlsx_rows(content: bytes) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        raise ValueError("openpyxl is not installed (required for .xlsx).")
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel sheet is empty.")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    if not any(headers):
        raise ValueError("Excel sheet header row is missing.")
    cols = { _norm_col(c): c for c in headers }
    dict_rows = []
    for r in rows[1:]:
        d = {}
        for idx, h in enumerate(headers):
            if not h:
                continue
            d[h] = r[idx] if idx < len(r) else None
        dict_rows.append(d)
    return _rows_from_dict_iter(dict_rows, cols)

def _rows_from_dict_iter(iterable, cols_norm_to_original: Dict[str, str]) -> List[Dict[str, Any]]:
    # Column mapping (case-insensitive + common variants)
    cert_keys = ["certificatenumber", "certno", "certnumber", "certificate", "cert"]
    forename_keys = ["forename", "firstname", "first_name", "givenname", "given_name"]
    surname_keys = ["surname", "lastname", "last_name"]
    issue_single_keys = ["issuedate", "dateofissue", "issue_date", "date_issued"]
    issue_dd_keys = ["issueday", "issue_day"]
    issue_mm_keys = ["issuemonth", "issue_month"]
    issue_yy_keys = ["issueyear", "issue_year"]
    dob_single_keys = ["dob", "dateofbirth", "dateofbirthddmmyyyy"]
    dd_keys = ["dobday", "day", "dd"]
    mm_keys = ["dobmonth", "month", "mm"]
    yy_keys = ["dobyear", "year", "yyyy", "yy"]

    def first_present(keys):
        for k in keys:
            if k in cols_norm_to_original:
                return cols_norm_to_original[k]
        return None

    cert_col = first_present(cert_keys)
    forename_col = first_present([_norm_col(k) for k in forename_keys])
    surname_col = first_present([_norm_col(k) for k in surname_keys])
    issue_col = first_present(issue_single_keys)
    issue_dd_col = first_present(issue_dd_keys)
    issue_mm_col = first_present(issue_mm_keys)
    issue_yy_col = first_present(issue_yy_keys)
    dob_col = first_present(dob_single_keys)
    dd_col = first_present(dd_keys)
    mm_col = first_present(mm_keys)
    yy_col = first_present(yy_keys)

    if not cert_col or not surname_col or not (dob_col or (dd_col and mm_col and yy_col)):
        raise ValueError("Spreadsheet must include Certificate No, Surname, and DOB columns (DOB or DOB Day/Month/Year).")

    out: List[Dict[str, Any]] = []
    for d in iterable:
        if not d:
            continue
        cert = str(d.get(cert_col) or "").strip()
        surname = str(d.get(surname_col) or "").strip()
        if dob_col:
            dd, mm, yy = _parse_dob_value(d.get(dob_col))
        else:
            dd = str(d.get(dd_col) or "").strip().zfill(2) if str(d.get(dd_col) or "").strip() else ""
            mm = str(d.get(mm_col) or "").strip().zfill(2) if str(d.get(mm_col) or "").strip() else ""
            yy = str(d.get(yy_col) or "").strip()
        # skip blank rows
        if not (cert or surname or (dd and mm and yy)):
            continue
        out.append({
            "certificate_number": cert,
            "surname": surname,
            "dob_day": dd,
            "dob_month": mm,
            "dob_year": yy,
        })
    return out
@app.post("/dbs/extract")
async def dbs_extract(files: List[UploadFile] = File(...)):
    if not files:
        raise HTTPException(status_code=400, detail="No file uploaded.")

    items: List[Dict[str, Any]] = []
    total_rows_cap = 100

    # Hard limit: 100 uploaded files (Premium)
    for file in files[:100]:
        content = await file.read()
        fname = (file.filename or "")
        filename = fname.lower()

        if len(content) > 25 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="File too large. Please upload files under 25MB.")

        
        # DOCX (best effort text extraction)
        if filename.endswith(".docx"):
            try:
                from docx import Document
            except Exception:
                raise HTTPException(status_code=400, detail="python-docx is not installed (required for .docx).")
            try:
                doc = Document(io.BytesIO(content))
                text = "\n".join([p.text for p in doc.paragraphs if (p.text or "").strip()])
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"DOCX '{fname}' could not be read: {str(e)}")
            fields = extract_fields_from_text(text)
            # no vision fallback for docx
            dob = fields.get("dob") if isinstance(fields.get("dob"), dict) else {}
            issue = fields.get("issue_date") if isinstance(fields.get("issue_date"), dict) else {}
            cert_val = fields.get("certificate_number") or ""
            surname_val = fields.get("surname") or ""
            dob_dd = (dob.get("dd") or "")
            dob_mm = (dob.get("mm") or "")
            dob_yy = (dob.get("yyyy") or "")
            issue_dd = (issue.get("dd") or "")
            issue_mm = (issue.get("mm") or "")
            issue_yy = (issue.get("yyyy") or "")
            source = {"certificate_number": "DOCX text" if cert_val else "", "surname": "DOCX text" if surname_val else "", "dob": "DOCX text" if (dob_dd and dob_mm and dob_yy) else "", "issue_date": "DOCX text" if (issue_dd and issue_mm and issue_yy) else ""}
            cert_conf = score_cert_number(cert_val)
            surname_conf = score_surname(surname_val, source=source.get("surname") or "")
            dob_conf = score_dob(dob_dd, dob_mm, dob_yy, source=source.get("dob") or "")
            issue_conf = score_dob(issue_dd, issue_mm, issue_yy, source=source.get("issue_date") or "") if (issue_dd or issue_mm or issue_yy) else 0
            items.append({
                "original_filename": fname,
                "forename": (fields.get("forename") or ""),
                "certificate_number": cert_val,
                "surname": surname_val,
                "dob_day": dob_dd,
                "dob_month": dob_mm,
                "dob_year": dob_yy,
                "issue_day": issue_dd,
                "issue_month": issue_mm,
                "issue_year": issue_yy,
                "confidence": {
                        "certificate_number": cert_conf,
                        "surname": surname_conf,
                        "dob": dob_conf,
                        "issue_date": issue_conf,
                    },
                "source": source,
            })
            if len(items) >= total_rows_cap:
                break
            continue

        # WEBP → convert to PNG for vision
        if filename.endswith(".webp"):
            try:
                from PIL import Image
                im = Image.open(io.BytesIO(content)).convert("RGB")
                buf = io.BytesIO()
                im.save(buf, format="PNG")
                content = buf.getvalue()
                filename = fname.lower().replace(".webp", ".png")
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"WEBP '{fname}' could not be processed: {str(e)}")

        # If spreadsheet: expand into rows (details already present)
        if filename.endswith(".csv") or filename.endswith(".xlsx"):
            try:
                if filename.endswith(".csv"):
                    rows = parse_csv_rows(content)
                else:
                    rows = parse_xlsx_rows(content)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Spreadsheet '{fname}' could not be read: {str(e)}")

            for r_idx, r in enumerate(rows, start=2):  # 2 = header is row 1
                if len(items) >= total_rows_cap:
                    break
                forename_val = (r.get("forename") or "").strip()
                cert_val = (r.get("certificate_number") or "").strip()
                surname_val = (r.get("surname") or "").strip()
                dob_dd = str(r.get("dob_day") or "").strip()
                dob_mm = str(r.get("dob_month") or "").strip()
                dob_yy = str(r.get("dob_year") or "").strip()
                issue_dd = str(r.get("issue_day") or "").strip()
                issue_mm = str(r.get("issue_month") or "").strip()
                issue_yy = str(r.get("issue_year") or "").strip()

                source = {"certificate_number": "Spreadsheet", "surname": "Spreadsheet", "dob": "Spreadsheet", "issue_date": "Spreadsheet", "forename": "Spreadsheet"}

                cert_conf = score_cert_number(cert_val)
                surname_conf = score_surname(surname_val, source="Spreadsheet")
                dob_conf = score_dob(dob_dd, dob_mm, dob_yy, source="Spreadsheet")

                items.append({
                    "original_filename": f"{fname} (Row {r_idx})",
                    "forename": forename_val,
                    "certificate_number": cert_val,
                    "surname": surname_val,
                    "dob_day": dob_dd,
                    "dob_month": dob_mm,
                    "dob_year": dob_yy,
                    "issue_day": issue_dd,
                    "issue_month": issue_mm,
                    "issue_year": issue_yy,
                    "confidence": {
                        "certificate_number": cert_conf,
                        "surname": surname_conf,
                        "dob": dob_conf,
                        "issue_date": score_dob(issue_dd, issue_mm, issue_yy, source="Spreadsheet") if (issue_dd or issue_mm or issue_yy) else 0,
                    },
                    "source": source,
                })

            if len(items) >= total_rows_cap:
                break
            continue

        # Otherwise: treat as PDF/image (extract)
        # Premium v2: Issue Date is extracted best-effort but hidden in UI (export only)
        fields: Dict[str, Any] = {"certificate_number": None, "surname": None, "dob": None, "issue_date": None}
        source: Dict[str, str] = {"certificate_number": "", "surname": "", "dob": "", "issue_date": ""}

        if filename.endswith(".pdf"):
            text = extract_text_from_pdf(content)
            if text and len(text) > 60:
                fields = extract_fields_from_text(text)
                if fields.get("certificate_number"):
                    source["certificate_number"] = "PDF text"
                if fields.get("surname"):
                    source["surname"] = "PDF text"
                if fields.get("dob"):
                    source["dob"] = "PDF text"
                if fields.get("issue_date"):
                    source["issue_date"] = "PDF text"

            if (not fields.get("certificate_number") or not fields.get("surname") or not fields.get("dob")):
                try:
                    imgs = pdf_to_images_bytes(content, max_pages=1, dpi=240)
                except Exception as e:
                    raise HTTPException(status_code=400, detail=str(e))
                images = [(b, "image/png") for b in imgs]
                vision = gemini_vision_extract_images(images)
                for k in ["certificate_number", "surname", "forename", "dob", "issue_date"]:
                    if not fields.get(k) and vision.get(k):
                        fields[k] = vision.get(k)
                        source[k] = "Image scan"
        else:
            is_png = content[:8] == b"\x89PNG\r\n\x1a\n"
            mime = "image/png" if is_png else "image/jpeg"
            vision = gemini_vision_extract_images([(content, mime)])
            fields = {
                "certificate_number": vision.get("certificate_number"),
                "surname": vision.get("surname"),
                "forename": vision.get("forename"),
                "dob": vision.get("dob"),
                "issue_date": vision.get("issue_date"),
            }
            for k in ["certificate_number", "surname", "dob", "issue_date"]:
                if fields.get(k):
                    source[k] = "Image scan"

        dob = fields.get("dob") if isinstance(fields.get("dob"), dict) else {}
        issue = fields.get("issue_date") if isinstance(fields.get("issue_date"), dict) else {}

        cert_val = fields.get("certificate_number") or ""
        surname_val = fields.get("surname") or ""
        dob_dd = (dob.get("dd") or "")
        dob_mm = (dob.get("mm") or "")
        dob_yy = (dob.get("yyyy") or "")
        issue_dd = (issue.get("dd") or "")
        issue_mm = (issue.get("mm") or "")
        issue_yy = (issue.get("yyyy") or "")

        # Prefer Gemini confidence when available (Image scan)
        vconf = (vision.get("confidence") if isinstance(locals().get("vision", {}), dict) else {}) or {}
        cert_conf = int(vconf.get("certificate_number") or score_cert_number(cert_val) or 0)
        surname_conf = int(vconf.get("surname") or score_surname(surname_val, source=source.get("surname") or "") or 0)
        dob_conf = int(vconf.get("dob") or score_dob(dob_dd, dob_mm, dob_yy, source=source.get("dob") or "") or 0)
        issue_conf = int(vconf.get("issue_date") or score_dob(issue_dd, issue_mm, issue_yy, source=source.get("issue_date") or "") or 0)

        items.append({
            "original_filename": fname,
            "forename": (fields.get("forename") or ""),
            "certificate_number": cert_val,
            "surname": surname_val,
            "dob_day": dob_dd,
            "dob_month": dob_mm,
            "dob_year": dob_yy,
            "issue_day": issue_dd,
            "issue_month": issue_mm,
            "issue_year": issue_yy,
            "confidence": {
                "certificate_number": cert_conf,
                "surname": surname_conf,
                "dob": dob_conf,
                # We intentionally do NOT return an overall confidence score.
                # The UI shows per-field confidence only.
                "issue_date": score_dob(issue_dd, issue_mm, issue_yy, source="Spreadsheet") if (issue_dd or issue_mm or issue_yy) else 0,
            },
            "source": source,
        })

        if len(items) >= total_rows_cap:
            break

    if len(items) >= total_rows_cap and len(files) > 0:
        # Soft notice (UI can show it)
        return JSONResponse({"items": items, "notice": "Row limit reached (100). Extra rows were skipped."})

    return JSONResponse({"items": items})


# -------------------------
# Optional exports (no extra API usage)
# -------------------------
def _dmy(dd: str, mm: str, yy: str) -> str:
    dd = str(dd or "").strip()
    mm = str(mm or "").strip()
    yy = str(yy or "").strip()
    if dd and mm and yy:
        return f"{dd.zfill(2)}/{mm.zfill(2)}/{yy}"
    return ""

    return "UNKNOWN"

def _export_rows_extract(items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out = []
    for it in (items or []):
        out.append({
            "Forename": (it.get("forename") or "").strip(),
            "Surname": (it.get("surname") or "").strip(),
            "Certificate Number": (it.get("certificate_number") or "").strip(),
            "DOB": _dmy(it.get("dob_day"), it.get("dob_month"), it.get("dob_year")),
            "Issue Date": _dmy(it.get("issue_day"), it.get("issue_month"), it.get("issue_year")),
            "PDF Filename": (it.get("original_filename") or "").strip(),
            "Notes": "",
        })
    return out

def _export_rows_results(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    checked_date = (payload.get("checked_date") or "").strip()
    rows = payload.get("rows") or []
    out = []
    for r in rows:
        status = (r.get("status") or "").strip()
        out.append({
            "Forename": (r.get("forename") or "").strip(),
            "Surname": (r.get("surname") or "").strip(),
            "Certificate Number": (r.get("certificate_number") or "").strip(),
            "DOB": _dmy(r.get("dob_day"), r.get("dob_month"), r.get("dob_year")),
            "Issue Date": _dmy(r.get("issue_day"), r.get("issue_month"), r.get("issue_year")),
            "Status": status,
                        "Checked Date": checked_date,
            "PDF Filename": (r.get("pdf_filename") or "").strip(),
            "Notes": (r.get("error") or r.get("notes") or "").strip(),
        })
    return out

def _csv_bytes(rows: List[Dict[str, Any]], columns: List[str]) -> bytes:
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
    w.writeheader()
    for r in rows:
        w.writerow({c: (r.get(c) if r.get(c) is not None else "") for c in columns})
    return buf.getvalue().encode("utf-8-sig")

def _xlsx_bytes(rows: List[Dict[str, Any]], columns: List[str]) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for r in rows:
        ws.append([r.get(c, "") for c in columns])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

@app.post("/dbs/export/extract")
async def export_extract(request: Request):
    data = await request.json()
    fmt = (data.get("format") or "xlsx").lower()
    items = data.get("items") or []
    rows = _export_rows_extract(items)
    columns = ["Forename","Surname","Certificate Number","DOB","Issue Date","PDF Filename","Notes"]
    if fmt == "csv":
        b = _csv_bytes(rows, columns)
        return Response(content=b, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=extract.csv"})
    else:
        b = _xlsx_bytes(rows, columns)
        return Response(content=b, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment; filename=extract.xlsx"})

@app.post("/dbs/export/results")
async def export_results(request: Request):
    data = await request.json()
    fmt = (data.get("format") or "xlsx").lower()
    rows = _export_rows_results(data)
    columns = ["Forename","Surname","Certificate Number","DOB","Issue Date","Status","Checked Date","PDF Filename","Notes"]
    if fmt == "csv":
        b = _csv_bytes(rows, columns)
        return Response(content=b, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=results.csv"})
    else:
        b = _xlsx_bytes(rows, columns)
        return Response(content=b, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": "attachment; filename=results.xlsx"})


# -------------------------

# Run DBS (single returns PDF; bulk returns job + links/zip)
# -------------------------
from dbs_runner import run_dbs_check_and_download_pdf


@app.get("/dbs/download/{job_id}/{name}")
async def dbs_download(job_id: str, name: str):
    meta = JOBS.get(job_id)
    if not meta:
        raise HTTPException(status_code=404, detail="Download expired or not found.")
    _touch_job(job_id)

    job_dir = Path(meta["path"])
    file_path = job_dir / name
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found.")

    mode = meta.get("mode") or ""

    # Cleanup rules (v2):
    # - Single: delete after download
    # - Bulk: if ZIP downloaded -> delete immediately; otherwise delete after 15 mins inactivity
    background = None
    if mode == "single":
        background = BackgroundTask(_delete_job, job_id)
    elif mode == "bulk" and name.lower().endswith(".zip"):
        meta["zip_downloaded"] = True
        background = BackgroundTask(_delete_job, job_id)

    return FileResponse(
        path=str(file_path),
        filename=name,
        media_type="application/octet-stream",
        background=background,
    )



@app.get("/dbs/status/{job_id}")
async def dbs_status(job_id: str):
    meta = JOBS.get(job_id)
    if not meta:
        raise HTTPException(status_code=404, detail="Job expired or not found.")
    _touch_job(job_id)
    rows = meta.get("rows") or []
    done = sum(1 for r in rows if (r.get("status") in ["clear", "needs_review", "portal_unavailable"]))
    total = len(rows) if rows else (1 if meta.get("mode") == "single" else 0)
    payload = {
        "job_id": job_id,
        "mode": meta.get("mode"),
        "state": meta.get("state"),
        "checked_date": meta.get("checked_date") or "",
        "running": {"done": done, "total": total},
        "rows": rows,
        "zip_ready": bool(meta.get("zip_ready")),
        "zip_name": meta.get("zip_name") or "",
        "zip_url": (f"/dbs/download/{job_id}/{meta.get('zip_name')}" if meta.get("zip_ready") and meta.get("zip_name") else ""),
        "message": meta.get("message") or "",
    }
    return JSONResponse(payload)


async def _process_bulk_job(
    *,
    job_id: str,
    job_dir: Path,
    organisation_name: str,
    employee_forename: str,
    employee_surname: str,
    items: List[Dict[str, Any]],
    previous_job_id: str = "",
):
    meta = JOBS.get(job_id)
    if not meta:
        return

    checked_date = _uk_checked_date()
    meta["checked_date"] = checked_date

    pdf_names: List[str] = []

    for i, it in enumerate(items[:100], start=1):
        meta = JOBS.get(job_id)
        if not meta:
            return

        row = meta["rows"][i - 1]
        row["error"] = ""
        _touch_job(job_id)

        certificate_number = (it.get("certificate_number") or "").strip()
        surname_extracted = (it.get("surname") or it.get("surname_extracted") or "").strip()
        surname_user = (it.get("surname_user") or "").strip()
        applicant_surname = (surname_extracted or surname_user).strip()

        dob_day = str(it.get("dob_day") or "").strip()
        dob_month = str(it.get("dob_month") or "").strip()
        dob_year = str(it.get("dob_year") or "").strip()

        
        # Option 1 (locked): if row is NOT dirty and we have previous job outputs, reuse/copy the previous PDF
        dirty = bool(it.get("dirty"))
        existing_status = (it.get("existing_status") or "").strip() or row.get("existing_status","").strip()
        existing_pdf = (it.get("existing_pdf_filename") or "").strip() or row.get("existing_pdf_filename","").strip()

        if (not dirty) and previous_job_id and existing_pdf:
            prev_meta = JOBS.get(previous_job_id)
            prev_dir = Path(prev_meta.get("path")) if prev_meta and prev_meta.get("path") else None
            prev_path = (prev_dir / existing_pdf) if prev_dir else None
            # Normalize status to supported set
            st0 = (existing_status or row.get("status") or "").strip()
            if st0 not in ["clear", "needs_review", "portal_unavailable"]:
                st0 = "needs_review"
            row["status"] = st0
            if st0 == "portal_unavailable":
                # No output allowed
                row["pdf_filename"] = ""
                row["pdf_url"] = ""
                row["error"] = "DBS portal unavailable (maintenance). Try later."
                continue

            if prev_path and prev_path.exists():
                out_surname = _safe_filename(applicant_surname.upper(), f"SURNAME{i}")
                cert_tmp = validate_cert_number(certificate_number)
                cert_part = _safe_filename(cert_tmp or certificate_number, f"CERT{i}")
                status_label = "Clear" if st0 == "clear" else ("Needs Review" if st0 == "needs_review" else "Portal Unavailable")
                final_name = _safe_filename(f"{out_surname} - {cert_part} - {status_label} - {checked_date}.pdf", f"DBS-Result-{i}.pdf")

                final_path = job_dir / final_name
                try:
                    if final_path.exists():
                        final_path.unlink()
                    import shutil
                    shutil.copyfile(str(prev_path), str(final_path))
                except Exception:
                    # fallback: keep name from previous
                    final_path = prev_path
                    final_name = existing_pdf

                pdf_names.append(final_name)
                row["pdf_filename"] = final_name
                row["filename"] = final_name
                row["pdf_url"] = f"/dbs/download/{job_id}/{final_name}"
                continue
        # Mark running only for rows that will actually be processed (dirty or no reusable output)
        row["status"] = "running"
        cert = validate_cert_number(certificate_number)

        if not cert:
            row["status"] = "needs_review"
            row["error"] = "Invalid certificate number."
            continue
        if not applicant_surname:
            row["status"] = "needs_review"
            row["error"] = "Applicant surname is missing."
            continue
        if not (dob_day and dob_month and dob_year):
            row["status"] = "needs_review"
            row["error"] = "DOB is incomplete."
            continue

        # Per-candidate subdir to avoid collisions within job
        cand_dir = job_dir / f"c{i:02d}"
        cand_dir.mkdir(parents=True, exist_ok=True)

        result = await anyio.to_thread.run_sync(
            lambda: run_dbs_check_and_download_pdf(
                organisation_name=organisation_name,
                employee_forename=employee_forename,
                employee_surname=employee_surname,
                certificate_number=cert,
                applicant_surname=applicant_surname,
                dob_day=str(dob_day).zfill(2),
                dob_month=str(dob_month).zfill(2),
                dob_year=str(dob_year),
                out_dir=cand_dir,
                headless=True,
            )
        )

        status = (result.get("status") or "needs_review").strip()
        if status not in ["clear", "needs_review", "portal_unavailable"]:
            status = "needs_review"

        row["status"] = status

        if status == "portal_unavailable":
            row["error"] = "DBS portal unavailable (maintenance). Try later."
            continue

        out_surname = _safe_filename(applicant_surname.upper(), f"SURNAME{i}")
        cert_part = _safe_filename(cert, f"CERT{i}")
        status_label = "Clear" if status == "clear" else ("Needs Review" if status == "needs_review" else "Portal Unavailable")
        final_name = f"{out_surname} - {cert_part} - {status_label} - {checked_date}.pdf"

        final_name = _safe_filename(final_name, f"DBS-Result-{i}.pdf")

        pdf_path = result.get("pdf_path") or ""
        if not pdf_path or not os.path.exists(pdf_path):
            # Some Needs Review cases are validation-form pages where we intentionally do not generate a PDF.
            if status == "needs_review" and result.get("no_pdf"):
                row["pdf_filename"] = ""
                row["pdf_url"] = ""
                row["notes"] = "Needs Review (no PDF available for validation page)."
                continue
            row["status"] = "needs_review"
            row["error"] = result.get("error") or "Runner did not produce a PDF."
            continue

        # Move into job root with final filename
        final_path = job_dir / final_name
        try:
            if final_path.exists():
                final_path.unlink()
            os.replace(pdf_path, str(final_path))
        except Exception:
            final_path = Path(pdf_path)
            final_name = final_path.name

        pdf_names.append(final_name)
        row["pdf_filename"] = final_name
        row["filename"] = final_name
        row["pdf_url"] = f"/dbs/download/{job_id}/{final_name}"

    # Create ZIP when all rows complete (exclude portal unavailable)
    meta = JOBS.get(job_id)
    if not meta:
        return

    if len(pdf_names) >= 2:
        zip_name = _safe_filename(f"DBS_Checks_{checked_date}.zip", "DBS_Checks.zip")
        zip_path = job_dir / zip_name
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for pdfn in pdf_names:
                fp = job_dir / pdfn
                if fp.exists():
                    zf.write(fp, arcname=pdfn)

        meta["zip_name"] = zip_name
        meta["zip_ready"] = True
        meta["message"] = ""
    elif len(pdf_names) == 1:
        meta["zip_name"] = ""
        meta["zip_ready"] = False
        meta["message"] = ""
    else:
        # Professional: do not offer an empty ZIP
        meta["zip_name"] = ""
        meta["zip_ready"] = False
        meta["message"] = "No PDFs available because the DBS portal is unavailable. Please try again later."

    meta["state"] = "done"
    _release_lock(job_id)
    _touch_job(job_id)


@app.post("/dbs/run")
async def dbs_run(request: Request):
    payload: Dict[str, Any] = {}
    ctype = (request.headers.get("content-type") or "").lower()
    if "application/json" in ctype:
        try:
            payload = await request.json()
        except Exception:
            payload = {}
    else:
        form = await request.form()
        payload = dict(form)

    sid = _sid_from(request)

    # Run lock: only 1 active running job per session
    existing = ACTIVE_BY_SID.get(sid)
    if existing:
        ex_meta = JOBS.get(existing)
        if ex_meta and (ex_meta.get("state") == "running"):
            raise HTTPException(status_code=409, detail="A job is already running for this session. Please wait for it to complete.")

    organisation_name = (payload.get("organisation_name") or payload.get("org_name") or payload.get("organisation") or "").strip()
    employee_forename = (payload.get("employee_forename") or payload.get("forename") or "").strip()
    employee_surname  = (payload.get("employee_surname")  or payload.get("surname_user") or payload.get("surname") or "").strip()
    previous_job_id = (payload.get("previous_job_id") or "").strip()


    if not (organisation_name and employee_forename and employee_surname):
        raise HTTPException(status_code=400, detail="Organisation/Forename/Surname (Step 1) is incomplete.")

    items = payload.get("items")
    if isinstance(items, list) and len(items) > 0:
        # Bulk mode: start job immediately and stream progress via /dbs/status polling
        job_id, job_dir = _new_job_dir(prefix="dbs", sid=sid, mode="bulk")
        meta = JOBS[job_id]
        meta["rows"] = []
        for i in range(min(100, len(items))):
            it = items[i] if isinstance(items[i], dict) else {}
            it0 = it
            dirty0 = bool(it0.get("dirty"))
            ex_status0 = (it0.get("existing_status") or "").strip()
            ex_pdf0 = (it0.get("existing_pdf_filename") or "").strip()
            # If rerun and row is not dirty, keep previous status/output visible immediately
            init_status = "queued"
            init_pdf_filename = ""
            init_pdf_url = ""
            if previous_job_id and (not dirty0) and ex_pdf0:
                st0 = ex_status0 if ex_status0 else "needs_review"
                if st0 not in ["clear","needs_review","portal_unavailable"]:
                    st0 = "needs_review"
                init_status = st0
                init_pdf_filename = ex_pdf0 if st0 != "portal_unavailable" else ""
                init_pdf_url = (f"/dbs/download/{previous_job_id}/{ex_pdf0}" if init_pdf_filename else "")
            meta["rows"].append({
                "row": i + 1,
                "status": init_status,
                "certificate_number": (it.get("certificate_number") or ""),
                "surname": (it.get("surname") or ""),
                "forename": (it.get("forename") or ""),
                "dob_day": (it.get("dob_day") or ""),
                "dob_month": (it.get("dob_month") or ""),
                "dob_year": (it.get("dob_year") or ""),
                "issue_day": (it.get("issue_day") or ""),
                "issue_month": (it.get("issue_month") or ""),
                "issue_year": (it.get("issue_year") or ""),
                "original_filename": (it.get("original_filename") or ""),
                "dirty": bool(it.get("dirty")),
                "existing_status": (it.get("existing_status") or ""),
                "existing_pdf_filename": (it.get("existing_pdf_filename") or ""),
                "pdf_filename": init_pdf_filename,
                "pdf_url": init_pdf_url,
                "error": "",
            })

        meta["state"] = "running"

        import asyncio
        asyncio.get_event_loop().create_task(
            _process_bulk_job(
                job_id=job_id,
                job_dir=job_dir,
                organisation_name=organisation_name,
                employee_forename=employee_forename,
                employee_surname=employee_surname,
                items=items,
                previous_job_id=previous_job_id,
            )
        )

        return JSONResponse({
            "job_id": job_id,
            "mode": "bulk",
            "status_url": f"/dbs/status/{job_id}",
            "rows": meta["rows"],
        })

    # Single mode
    certificate_number = (payload.get("certificate_number") or "").strip()
    surname_user = (payload.get("surname_user") or payload.get("surname") or "").strip()
    surname_extracted = (payload.get("surname_extracted") or "").strip()
    applicant_surname = (surname_extracted or surname_user).strip()

    dob_day = (payload.get("dob_day") or payload.get("dob_dd") or "").strip()
    dob_month = (payload.get("dob_month") or payload.get("dob_mm") or "").strip()
    dob_year = (payload.get("dob_year") or payload.get("dob_yyyy") or "").strip()

    cert = validate_cert_number(certificate_number)
    if not cert:
        raise HTTPException(status_code=400, detail="Invalid certificate number.")
    if not applicant_surname:
        raise HTTPException(status_code=400, detail="Applicant surname is missing.")
    if not (dob_day and dob_month and dob_year):
        raise HTTPException(status_code=400, detail="DOB is incomplete.")

    job_id, job_dir = _new_job_dir(prefix="dbs", sid=sid, mode="single")
    checked_date = _uk_checked_date()
    JOBS[job_id]["checked_date"] = checked_date

    result = await anyio.to_thread.run_sync(
        lambda: run_dbs_check_and_download_pdf(
            organisation_name=organisation_name,
            employee_forename=employee_forename,
            employee_surname=employee_surname,
            certificate_number=cert,
            applicant_surname=applicant_surname,
            dob_day=str(dob_day).zfill(2),
            dob_month=str(dob_month).zfill(2),
            dob_year=str(dob_year),
            out_dir=job_dir,
            headless=True,
        )
    )

    status = (result.get("status") or "needs_review").strip()
    if status not in ["clear", "needs_review", "portal_unavailable"]:
        status = "needs_review"

    if status == "portal_unavailable":
        # no PDF generated
        JOBS[job_id]["state"] = "done"
        JOBS[job_id]["message"] = "DBS portal unavailable (maintenance). Try later."
        _release_lock(job_id)
        _delete_job(job_id)
        return JSONResponse({
            "ok": True,
            "status": "portal_unavailable",
            "message": "DBS portal unavailable (maintenance). Try later.",
        })

    out_surname = _safe_filename(applicant_surname.upper(), "SURNAME")

    cert_part = _safe_filename(certificate_number, "CERT")
    status_label = "Clear" if status == "clear" else ("Needs Review" if status == "needs_review" else "Portal Unavailable")
    final_name = f"{out_surname} - {cert_part} - {status_label} - {checked_date}.pdf"

    final_name = _safe_filename(final_name, "DBS-Result.pdf")

    pdf_path = result.get("pdf_path") or ""
    if not pdf_path or not os.path.exists(pdf_path):
        JOBS[job_id]["state"] = "done"
        _release_lock(job_id)
        raise HTTPException(status_code=500, detail="Runner succeeded but PDF was not found.")

    final_path = job_dir / final_name
    try:
        if final_path.exists():
            final_path.unlink()
        os.replace(pdf_path, str(final_path))
    except Exception:
        final_name = Path(pdf_path).name

    JOBS[job_id]["state"] = "done"
    _release_lock(job_id)

    return JSONResponse({
        "ok": True,
        "status": status,
        "job_id": job_id,
        "checked_date": checked_date,
        "filename": final_name,
        "pdf_url": f"/dbs/download/{job_id}/{final_name}",
    })